"""Validation utility functions."""

import re
import zipfile
from pathlib import Path
from typing import Optional
from zipfile import BadZipFile

from docx import Document
from openpyxl import load_workbook

from app.core.exceptions import (
    ValidationError,
    InvalidVulnerabilityFormatError,
    MissingRequiredColumnError,
)
from app.core.logging import get_logger

logger = get_logger(__name__)


def validate_docx_file(file_path: Path) -> None:
    """
    Validate that file is a valid Word document.

    Args:
        file_path: Path to .docx file

    Raises:
        ValidationError: If file is not a valid Word document
    """
    if not file_path.exists():
        raise ValidationError(f"File not found: {file_path}")

    if file_path.suffix.lower() != ".docx":
        raise ValidationError(f"File must be a .docx file, got: {file_path.suffix}")

    try:
        # Attempt to open document
        Document(str(file_path))
        logger.info(f"Valid Word document: {file_path}")
    except BadZipFile:
        raise ValidationError(
            f"File is not a valid Word document (corrupted or wrong format): {file_path}"
        )
    except Exception as e:
        raise ValidationError(f"Failed to validate Word document: {str(e)}")


def validate_excel_file(
    file_path: Path, required_columns: Optional[list[str]] = None
) -> None:
    """
    Validate that file is a valid Excel file with required columns.

    Args:
        file_path: Path to .xlsx file
        required_columns: List of required column names

    Raises:
        ValidationError: If file is not valid
        MissingRequiredColumnError: If required columns are missing
    """
    if not file_path.exists():
        raise ValidationError(f"File not found: {file_path}")

    if file_path.suffix.lower() not in [".xlsx", ".xls"]:
        raise ValidationError(f"File must be an Excel file, got: {file_path.suffix}")

    try:
        # Load workbook
        workbook = load_workbook(str(file_path), read_only=True)

        if not workbook.sheetnames:
            raise ValidationError("Excel file has no sheets")

        # Get first sheet
        sheet = workbook[workbook.sheetnames[0]]

        # Check for required columns if specified
        if required_columns:
            # Get header row (first row)
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

            missing_columns = [col for col in required_columns if col not in headers]

            if missing_columns:
                raise MissingRequiredColumnError(
                    f"Missing required columns: {', '.join(missing_columns)}",
                    details={
                        "missing_columns": missing_columns,
                        "found_columns": headers,
                    },
                )

        workbook.close()
        logger.info(f"Valid Excel file: {file_path}")

    except MissingRequiredColumnError:
        raise
    except Exception as e:
        raise ValidationError(f"Failed to validate Excel file: {str(e)}")


def validate_poc_folder(folder_path: str) -> Path:
    """
    Validate that PoC folder exists and is accessible.

    Args:
        folder_path: Path to PoC folder

    Returns:
        Validated Path object

    Raises:
        ValidationError: If folder is invalid
    """
    path = Path(folder_path)

    # Prevent path traversal
    try:
        path = path.resolve()
    except Exception as e:
        raise ValidationError(f"Invalid path: {str(e)}")

    if not path.exists():
        logger.warning(f"PoC folder does not exist: {path}")
        raise ValidationError(f"PoC folder not found: {path}")

    if not path.is_dir():
        raise ValidationError(f"Path is not a directory: {path}")

    return path


def validate_zip_file(file_path: Path) -> None:
    """
    Validate that file is a valid ZIP file.

    Args:
        file_path: Path to .zip file

    Raises:
        ValidationError: If file is not a valid ZIP file
    """
    if not file_path.exists():
        raise ValidationError(f"File not found: {file_path}")

    if file_path.suffix.lower() != ".zip":
        raise ValidationError(f"File must be a .zip file, got: {file_path.suffix}")

    try:
        # Check if it's a valid ZIP file
        if not zipfile.is_zipfile(file_path):
            raise ValidationError(f"File is not a valid ZIP file: {file_path}")

        # Try to open it
        with zipfile.ZipFile(file_path, "r") as zip_ref:
            # Test the ZIP integrity
            zip_ref.testzip()

        logger.info(f"Valid ZIP file: {file_path}")

    except BadZipFile:
        raise ValidationError(
            f"File is not a valid ZIP file (corrupted or wrong format): {file_path}"
        )
    except Exception as e:
        raise ValidationError(f"Failed to validate ZIP file: {str(e)}")


def validate_vulnerability_id(vuln_id: str) -> str:
    """
    Validate vulnerability ID format (H1, M1, L1, etc.).

    Args:
        vuln_id: Vulnerability ID string

    Returns:
        Validated and normalized vulnerability ID

    Raises:
        InvalidVulnerabilityFormatError: If format is invalid
    """
    if not vuln_id:
        raise InvalidVulnerabilityFormatError("Vulnerability ID cannot be empty")

    # Pattern: starts with C/H/M/L/I followed by digits
    pattern = r"^[CHMLI]\d+$"

    vuln_id_upper = vuln_id.strip().upper()

    if not re.match(pattern, vuln_id_upper):
        raise InvalidVulnerabilityFormatError(
            f"Invalid vulnerability ID format: '{vuln_id}'. "
            "Expected format: [C|H|M|L|I]<number> (e.g., H1, M2, L3)",
            details={"provided": vuln_id, "pattern": pattern},
        )

    return vuln_id_upper
