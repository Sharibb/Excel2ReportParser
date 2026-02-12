"""Excel reader for ingesting structured vulnerability data."""

from pathlib import Path
from typing import List, Dict, Any

from openpyxl import load_workbook

from app.core.exceptions import ParsingError, MissingRequiredColumnError
from app.core.logging import get_logger
from app.models.vulnerability import (
    Vulnerability,
    VulnerabilityExcelRow,
    VulnerabilityReport,
)
from app.utils.validators import validate_excel_file

logger = get_logger(__name__)


class ExcelReader:
    """Reader for parsing vulnerability data from Excel files."""

    REQUIRED_COLUMNS = [
        "Vulnerability ID",
        "Title",
        "Description",
        "Risk Level",
        "Affected Components",
        "Recommendation",
    ]

    def __init__(self, excel_path: Path) -> None:
        """
        Initialize Excel reader.

        Args:
            excel_path: Path to Excel file
        """
        self.excel_path = excel_path

    def read(self) -> VulnerabilityReport:
        """
        Read Excel file and parse vulnerability data.

        Returns:
            VulnerabilityReport containing parsed data

        Raises:
            ParsingError: If parsing fails
            MissingRequiredColumnError: If required columns are missing
        """
        try:
            # Validate file and required columns
            validate_excel_file(self.excel_path, self.REQUIRED_COLUMNS)

            # Load workbook
            workbook = load_workbook(str(self.excel_path), read_only=True)
            sheet = workbook[workbook.sheetnames[0]]

            # Get headers
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

            # Parse rows
            vulnerabilities: List[Vulnerability] = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                try:
                    # Create dictionary from row
                    row_data = dict(zip(headers, row))

                    # Skip empty rows
                    if not row_data.get("Vulnerability ID"):
                        continue

                    # Parse vulnerability
                    excel_row = VulnerabilityExcelRow(**row_data)
                    vulnerability = excel_row.to_vulnerability()
                    vulnerabilities.append(vulnerability)

                except Exception as e:
                    logger.warning(f"Failed to parse row: {e}")
                    continue

            workbook.close()

            # Create report
            report = VulnerabilityReport(
                report_title="Vulnerability Report",
                vulnerabilities=vulnerabilities,
            )
            report.calculate_counts()

            logger.info(
                f"Successfully read {len(vulnerabilities)} vulnerabilities "
                f"from {self.excel_path}"
            )

            return report

        except MissingRequiredColumnError:
            raise
        except Exception as e:
            logger.error(f"Failed to read Excel file: {e}")
            raise ParsingError(f"Failed to read Excel file: {str(e)}")

    def read_as_dicts(self) -> List[Dict[str, Any]]:
        """
        Read Excel file and return raw dictionaries.

        Returns:
            List of dictionaries with vulnerability data

        Raises:
            ParsingError: If parsing fails
        """
        try:
            validate_excel_file(self.excel_path, self.REQUIRED_COLUMNS)

            workbook = load_workbook(str(self.excel_path), read_only=True)
            sheet = workbook[workbook.sheetnames[0]]

            # Get headers
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

            # Parse rows as dictionaries
            rows: List[Dict[str, Any]] = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))

                # Skip empty rows
                if not row_data.get("Vulnerability ID"):
                    continue

                rows.append(row_data)

            workbook.close()

            logger.info(f"Successfully read {len(rows)} rows from {self.excel_path}")
            return rows

        except Exception as e:
            logger.error(f"Failed to read Excel as dicts: {e}")
            raise ParsingError(f"Failed to read Excel file: {str(e)}")
