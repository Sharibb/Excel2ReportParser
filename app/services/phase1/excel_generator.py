"""Excel generator for creating structured vulnerability reports."""

from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.core.exceptions import FileProcessingError
from app.core.logging import get_logger
from app.models.vulnerability import Vulnerability, VulnerabilityReport

logger = get_logger(__name__)


class ExcelGenerator:
    """Generator for creating Excel files from vulnerability data."""

    # Define column headers
    HEADERS = [
        "Vulnerability ID",
        "Title",
        "Description",
        "Risk Level",
        "CVSS Score",
        "Affected Components",
        "Recommendation",
        "POC_Folder",
        "Steps",
        "CWE ID",
        "Impact",
        "References",
        "Remediation Effort",
    ]

    def __init__(self) -> None:
        """Initialize Excel generator."""
        self.workbook: Workbook = Workbook()
        self.sheet: Worksheet = self.workbook.active
        self.sheet.title = "Vulnerabilities"

    def generate(
        self, report: VulnerabilityReport, output_path: Path
    ) -> Path:
        """
        Generate Excel file from vulnerability report.

        Args:
            report: VulnerabilityReport containing vulnerabilities
            output_path: Path where Excel file should be saved

        Returns:
            Path to generated Excel file

        Raises:
            FileProcessingError: If generation fails
        """
        try:
            # Create headers
            self._create_headers()

            # Add vulnerabilities
            self._add_vulnerabilities(report.vulnerabilities)

            # Apply formatting
            self._apply_formatting()

            # Save file
            output_path.parent.mkdir(parents=True, exist_ok=True)
            self.workbook.save(str(output_path))

            logger.info(
                f"Successfully generated Excel file with {len(report.vulnerabilities)} "
                f"vulnerabilities at {output_path}"
            )

            return output_path

        except Exception as e:
            logger.error(f"Failed to generate Excel file: {e}")
            raise FileProcessingError(f"Failed to generate Excel file: {str(e)}")

    def _create_headers(self) -> None:
        """Create and style header row."""
        # Write headers
        for col_idx, header in enumerate(self.HEADERS, start=1):
            cell = self.sheet.cell(row=1, column=col_idx, value=header)

            # Style header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid"
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Set row height
        self.sheet.row_dimensions[1].height = 30

    def _add_vulnerabilities(self, vulnerabilities: List[Vulnerability]) -> None:
        """
        Add vulnerability data rows.

        Args:
            vulnerabilities: List of Vulnerability objects
        """
        for row_idx, vuln in enumerate(vulnerabilities, start=2):
            # Basic fields
            self.sheet.cell(row=row_idx, column=1, value=vuln.vuln_id)
            self.sheet.cell(row=row_idx, column=2, value=vuln.title)
            self.sheet.cell(row=row_idx, column=3, value=vuln.description)
            self.sheet.cell(row=row_idx, column=4, value=vuln.risk_level)
            self.sheet.cell(row=row_idx, column=5, value=vuln.cvss_score)
            self.sheet.cell(row=row_idx, column=6, value=vuln.affected_components)
            self.sheet.cell(row=row_idx, column=7, value=vuln.recommendation)
            self.sheet.cell(row=row_idx, column=8, value=vuln.poc_folder)

            # PoC steps - join with semicolon delimiter
            if vuln.steps:
                steps_delimited = "; ".join(vuln.steps)
                self.sheet.cell(row=row_idx, column=9, value=steps_delimited)
            
            # Additional fields
            self.sheet.cell(row=row_idx, column=10, value=vuln.cwe_id)
            self.sheet.cell(row=row_idx, column=11, value=vuln.impact)
            self.sheet.cell(row=row_idx, column=12, value=vuln.references)
            self.sheet.cell(row=row_idx, column=13, value=vuln.remediation_effort)

    def _apply_formatting(self) -> None:
        """Apply formatting to the worksheet."""
        # Set column widths
        column_widths = {
            1: 15,  # Vulnerability ID
            2: 30,  # Title
            3: 50,  # Description
            4: 12,  # Risk Level
            5: 12,  # CVSS Score
            6: 30,  # Affected Components
            7: 50,  # Recommendation
            8: 20,  # POC_Folder
            9: 60,  # Steps (wider for delimited content)
            10: 15,  # CWE ID
            11: 40,  # Impact
            12: 40,  # References
            13: 15,  # Remediation Effort
        }

        for col_idx, width in column_widths.items():
            column_letter = self.sheet.cell(row=1, column=col_idx).column_letter
            self.sheet.column_dimensions[column_letter].width = width

        # Apply text wrapping to multi-line columns
        for row in self.sheet.iter_rows(min_row=2, max_col=13):
            # Description
            row[2].alignment = Alignment(wrap_text=True, vertical="top")
            # Recommendation
            row[6].alignment = Alignment(wrap_text=True, vertical="top")
            # Steps
            row[8].alignment = Alignment(wrap_text=True, vertical="top")
            # Impact
            if len(row) > 10:
                row[10].alignment = Alignment(wrap_text=True, vertical="top")

        # Freeze header row
        self.sheet.freeze_panes = "A2"

    @staticmethod
    def create_from_report(
        report: VulnerabilityReport, output_path: Path
    ) -> Path:
        """
        Convenience method to create Excel file from report.

        Args:
            report: VulnerabilityReport object
            output_path: Path where Excel file should be saved

        Returns:
            Path to generated Excel file
        """
        generator = ExcelGenerator()
        return generator.generate(report, output_path)
