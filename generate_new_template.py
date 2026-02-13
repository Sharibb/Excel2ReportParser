"""Generate new Excel template with single Steps column (semicolon-delimited)."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def create_template() -> None:
    """Create new Excel template with semicolon-delimited Steps column."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"

    # Define headers
    headers = [
        "Vulnerability ID",
        "Title",
        "Description",
        "Risk Level",
        "CVSS Score",
        "Affected Components",
        "Recommendation",
        "POC_Folder",
        "Steps",
        "CWE ID",
        "Impact",
        "References",
        "Remediation Effort",
    ]

    # Style header row
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 30

    # Add example rows for ALL risk levels
    examples = [
        {
            "Vulnerability ID": "C1",
            "Title": "SQL Injection in Login Form",
            "Description": "The login form is vulnerable to SQL injection attacks, allowing attackers to bypass authentication and gain unauthorized access to the database.",
            "Risk Level": "Critical",
            "CVSS Score": "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:H (Base Score: 9.8)",
            "Affected Components": "https://example.com/login",
            "Recommendation": "Use parameterized queries or prepared statements. Implement input validation and proper authentication mechanisms. Apply principle of least privilege to database accounts.",
            "POC_Folder": "C1",
            "Steps": "Navigate to login page; Enter ' OR '1'='1 in username field; Submit the form; Observe successful authentication bypass",
            "CWE ID": "CWE-89",
            "Impact": "Complete compromise of authentication system, unauthorized database access, potential data exfiltration",
            "References": "https://owasp.org/www-community/attacks/SQL_Injection",
            "Remediation Effort": "Medium",
        },
        {
            "Vulnerability ID": "C2",
            "Title": "Remote Code Execution via File Upload",
            "Description": "The application allows upload of executable files without proper validation, enabling attackers to execute arbitrary code on the server.",
            "Risk Level": "Critical",
            "CVSS Score": "CVSS:3.1/AV:N/AC:L/PR:L/UI:N/S:U/C:H/I:H/A:H (Base Score: 8.8)",
            "Affected Components": "https://example.com/upload",
            "Recommendation": "Implement strict file type validation. Store uploaded files outside webroot. Use whitelist-based file extension checking. Scan files for malware.",
            "POC_Folder": "C2",
            "Steps": "Login to application; Navigate to upload page; Upload malicious PHP file; Access uploaded file directly; Execute arbitrary commands",
            "CWE ID": "CWE-434",
            "Impact": "Full server compromise, data breach, service disruption",
            "References": "https://owasp.org/www-community/vulnerabilities/Unrestricted_File_Upload",
            "Remediation Effort": "High",
        },
        {
            "Vulnerability ID": "H1",
            "Title": "Cross-Site Scripting (XSS) in Search",
            "Description": "The application does not properly sanitize user input in the search functionality, allowing attackers to inject malicious scripts that execute in victims' browsers.",
            "Risk Level": "High",
            "CVSS Score": "CVSS:3.1/AV:N/AC:L/PR:N/UI:R/S:C/C:L/I:L/A:N (Base Score: 6.1)",
            "Affected Components": "https://example.com/search",
            "Recommendation": "Implement proper input validation and output encoding. Use Content Security Policy headers. Sanitize all user inputs before rendering.",
            "POC_Folder": "H1",
            "Steps": "Navigate to search page; Enter <script>alert('XSS')</script> in search box; Submit search; Observe script execution in browser",
            "CWE ID": "CWE-79",
            "Impact": "Session hijacking, credential theft, defacement",
            "References": "https://owasp.org/www-community/attacks/xss/",
            "Remediation Effort": "Low",
        },
        {
            "Vulnerability ID": "H2",
            "Title": "Broken Authentication - Weak Password Policy",
            "Description": "The application allows weak passwords and does not implement account lockout mechanisms, making brute force attacks feasible.",
            "Risk Level": "High",
            "CVSS Score": "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:H/I:H/A:N (Base Score: 9.1)",
            "Affected Components": "https://example.com/register, https://example.com/login",
            "Recommendation": "Enforce strong password policies (minimum 12 characters, complexity requirements). Implement account lockout after failed attempts. Add rate limiting and CAPTCHA.",
            "POC_Folder": "H2",
            "Steps": "Attempt to register with password '123'; Observe acceptance; Launch brute force attack; Observe no rate limiting or lockout",
            "CWE ID": "CWE-521",
            "Impact": "Account compromise, unauthorized access to user data",
            "References": "https://owasp.org/www-project-top-ten/2017/A2_2017-Broken_Authentication",
            "Remediation Effort": "Medium",
        },
        {
            "Vulnerability ID": "M1",
            "Title": "Insecure Direct Object Reference (IDOR)",
            "Description": "Users can access other users' data by manipulating URL parameters without proper authorization checks.",
            "Risk Level": "Medium",
            "CVSS Score": "CVSS:3.1/AV:N/AC:L/PR:L/UI:N/S:U/C:L/I:N/A:N (Base Score: 4.3)",
            "Affected Components": "https://example.com/profile?id=123",
            "Recommendation": "Implement proper authorization checks for all user-specific resources. Use indirect references or session-based access control. Validate user permissions server-side.",
            "POC_Folder": "M1",
            "Steps": "Login as user A with id=100; Navigate to profile page; Change id parameter to 101 in URL; Observe unauthorized access to user B's profile data",
            "CWE ID": "CWE-639",
            "Impact": "Unauthorized access to sensitive user data, privacy violation",
            "References": "https://owasp.org/www-project-top-ten/2017/A5_2017-Broken_Access_Control",
            "Remediation Effort": "Medium",
        },
        {
            "Vulnerability ID": "M2",
            "Title": "Missing Rate Limiting on API Endpoints",
            "Description": "API endpoints lack rate limiting controls, allowing potential denial of service or brute force attacks.",
            "Risk Level": "Medium",
            "CVSS Score": "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:N/I:N/A:L (Base Score: 5.3)",
            "Affected Components": "https://api.example.com/v1/*",
            "Recommendation": "Implement rate limiting on all API endpoints. Use token bucket or sliding window algorithms. Return appropriate 429 status codes when limits are exceeded.",
            "POC_Folder": "M2",
            "Steps": "Send 1000 requests to API endpoint; Observe no rate limiting; Monitor server resource consumption; Observe potential service degradation",
            "CWE ID": "CWE-770",
            "Impact": "Service degradation, potential denial of service",
            "References": "https://owasp.org/www-project-api-security/",
            "Remediation Effort": "Low",
        },
        {
            "Vulnerability ID": "L1",
            "Title": "Information Disclosure via Error Messages",
            "Description": "Detailed error messages expose sensitive information about the application's internal structure, database schema, and file paths.",
            "Risk Level": "Low",
            "CVSS Score": "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:L/I:N/A:N (Base Score: 5.3)",
            "Affected Components": "https://example.com/admin",
            "Recommendation": "Implement generic error messages for users. Log detailed errors server-side only. Disable debug mode in production. Use custom error pages.",
            "POC_Folder": "L1",
            "Steps": "Navigate to admin page; Enter invalid SQL syntax; Observe detailed database error message; Extract database version and table information",
            "CWE ID": "CWE-209",
            "Impact": "Information leakage aids further attacks, reveals system architecture",
            "References": "https://owasp.org/www-project-top-ten/2017/A3_2017-Sensitive_Data_Exposure",
            "Remediation Effort": "Low",
        },
        {
            "Vulnerability ID": "L2",
            "Title": "Missing Security Headers",
            "Description": "The application does not implement recommended security headers such as X-Frame-Options, X-Content-Type-Options, and Strict-Transport-Security.",
            "Risk Level": "Low",
            "CVSS Score": "CVSS:3.1/AV:N/AC:H/PR:N/UI:R/S:U/C:L/I:L/A:N (Base Score: 4.2)",
            "Affected Components": "All application pages",
            "Recommendation": "Implement security headers: X-Frame-Options: DENY, X-Content-Type-Options: nosniff, Strict-Transport-Security, Content-Security-Policy, X-XSS-Protection.",
            "POC_Folder": "L2",
            "Steps": "Inspect HTTP response headers; Verify absence of security headers; Attempt clickjacking attack; Observe lack of protection",
            "CWE ID": "CWE-693",
            "Impact": "Increased attack surface, potential for clickjacking and MIME sniffing attacks",
            "References": "https://owasp.org/www-project-secure-headers/",
            "Remediation Effort": "Low",
        },
        {
            "Vulnerability ID": "I1",
            "Title": "Verbose Server Banner",
            "Description": "The web server exposes detailed version information in HTTP headers and error pages, aiding attackers in reconnaissance.",
            "Risk Level": "Informational",
            "CVSS Score": "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:N/I:N/A:N (Base Score: 0.0)",
            "Affected Components": "Server: Apache/2.4.41 (Ubuntu)",
            "Recommendation": "Configure web server to suppress version information. Use ServerTokens Prod and ServerSignature Off in Apache. Remove or modify default error pages.",
            "POC_Folder": "I1",
            "Steps": "Send HTTP request to server; Inspect Server header in response; Note specific version information disclosed",
            "CWE ID": "CWE-200",
            "Impact": "Information disclosure aids targeted attacks, reveals potential vulnerabilities in specific versions",
            "References": "https://owasp.org/www-project-web-security-testing-guide/",
            "Remediation Effort": "Low",
        },
        {
            "Vulnerability ID": "I2",
            "Title": "Directory Listing Enabled",
            "Description": "Web server allows directory browsing on certain paths, exposing file structure and potentially sensitive files.",
            "Risk Level": "Informational",
            "CVSS Score": "CVSS:3.1/AV:N/AC:L/PR:N/UI:N/S:U/C:N/I:N/A:N (Base Score: 0.0)",
            "Affected Components": "https://example.com/uploads/, https://example.com/assets/",
            "Recommendation": "Disable directory listing in web server configuration. Add index files to directories. Configure Options -Indexes in Apache or autoindex off in Nginx.",
            "POC_Folder": "I2",
            "Steps": "Navigate to /uploads/ directory; Observe directory listing; Review exposed file names and structure",
            "CWE ID": "CWE-548",
            "Impact": "Exposure of file structure and names, potential discovery of sensitive files",
            "References": "https://cwe.mitre.org/data/definitions/548.html",
            "Remediation Effort": "Low",
        },
    ]

    # Add example data
    for row_idx, example in enumerate(examples, start=2):
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=example.get(header, ""))

    # Set column widths
    column_widths = {
        1: 15,  # Vulnerability ID
        2: 30,  # Title
        3: 50,  # Description
        4: 12,  # Risk Level
        5: 12,  # CVSS Score
        6: 30,  # Affected Components
        7: 50,  # Recommendation
        8: 20,  # POC_Folder
        9: 60,  # Steps (wider for delimited content)
        10: 15,  # CWE ID
        11: 40,  # Impact
        12: 40,  # References
        13: 15,  # Remediation Effort
    }

    for col_idx, width in column_widths.items():
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        ws.column_dimensions[column_letter].width = width

    # Apply text wrapping
    for row in ws.iter_rows(min_row=2, max_col=13):
        row[2].alignment = Alignment(wrap_text=True, vertical="top")  # Description
        row[6].alignment = Alignment(wrap_text=True, vertical="top")  # Recommendation
        row[8].alignment = Alignment(wrap_text=True, vertical="top")  # Steps
        if len(row) > 10:
            row[10].alignment = Alignment(wrap_text=True, vertical="top")  # Impact

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save template
    output_path = Path(__file__).parent / "All_Risk_Levels_Template.xlsx"
    
    # Try to save, if locked, try backup name
    try:
        wb.save(str(output_path))
        print(f"[SUCCESS] Created new template: {output_path}")
    except PermissionError:
        backup_path = Path(__file__).parent / "All_Risk_Levels_Template_NEW.xlsx"
        wb.save(str(backup_path))
        print(f"[WARNING] Original file is open/locked")
        print(f"[SUCCESS] Created template as: {backup_path}")
        print(f"[INFO] Close the original file and rename this to All_Risk_Levels_Template.xlsx")
        output_path = backup_path

    # Also save to frontend directory
    frontend_path = Path(__file__).parent / "frontend" / "All_Risk_Levels_Template.xlsx"
    if frontend_path.parent.exists():
        try:
            wb.save(str(frontend_path))
            print(f"[SUCCESS] Created frontend template: {frontend_path}")
        except PermissionError:
            frontend_backup = Path(__file__).parent / "frontend" / "All_Risk_Levels_Template_NEW.xlsx"
            wb.save(str(frontend_backup))
            print(f"[WARNING] Frontend file is open/locked")
            print(f"[SUCCESS] Created frontend template as: {frontend_backup}")


if __name__ == "__main__":
    create_template()
