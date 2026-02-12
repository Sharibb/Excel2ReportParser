import openpyxl
import sys

wb = openpyxl.load_workbook("All_Risk_Levels_Template.xlsx")
ws = wb.active

print("=== EXCEL STRUCTURE ===")
print(f"Sheet name: {ws.title}")
print(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
print()

# Print headers
print("=== HEADERS (Row 1) ===")
headers = []
for col in range(1, ws.max_column + 1):
    cell_value = ws.cell(1, col).value
    headers.append(cell_value)
    print(f"Column {col}: {cell_value}")

print()
print("=== FIRST DATA ROW (Row 2) ===")
for col in range(1, min(ws.max_column + 1, 12)):  # First 11 columns
    cell_value = ws.cell(2, col).value
    print(f"{headers[col-1]}: {cell_value}")

print()
print("=== CHECKING POC COLUMNS ===")
for col in range(12, ws.max_column + 1):
    header = ws.cell(1, col).value
    value = ws.cell(2, col).value
    if value:
        print(f"{header}: {value[:100] if isinstance(value, str) and len(value) > 100 else value}")
